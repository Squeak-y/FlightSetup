import tkinter as tk
import os
import webbrowser
import openpyxl
import re
import random
import subprocess
import threading
import configparser
import sys
from tkinter import ttk
from tkinter import font as tkFont

# Get the directory of the script or bundle
if getattr(sys, 'frozen', False):
    current_directory = os.path.dirname(sys.executable)
else:
    current_directory = os.path.dirname(os.path.abspath(__file__))

database_path = os.path.join(current_directory, "Database.xlsx")
config_path = os.path.join(current_directory, "config.ini")

def is_font_present(font_name):
    root = tk.Tk()
    available_fonts = list(tkFont.families())
    root.destroy()
    return font_name in available_fonts

desired_font = "Space Grotesk"
backup_font = "Segoe UI"
font_to_use = desired_font if is_font_present(desired_font) else backup_font
default_font = (font_to_use, 12, "bold")

def get_path_from_config(resource_name, default=None):
    config = configparser.ConfigParser()
    config.read(config_path)
    
    if 'Paths' not in config.sections():
        raise Exception("The 'Paths' section is not found in config.ini!")
    
    if resource_name not in config['Paths']:
        if default:
            return default
        raise Exception(f"'{resource_name}' not found in 'Paths' section of config.ini!")
    
    return os.path.join(current_directory, config['Paths'][resource_name])

def launch_volanta():
    volanta_path = get_path_from_config('volanta')
    subprocess.Popen(volanta_path, shell=True)

def on_button_click():
    thread = threading.Thread(target=launch_volanta)
    thread.start()

def open_path(path):
    webbrowser.open(path)

def get_airline_code(flight_number):
    match = re.match(r'([A-Za-z]+)', flight_number)
    return match.group(0).lower() if match else ""

def reset_fields():
    for entry in entries.values():
        entry.delete(0, tk.END)
    
    aircraft_dropdown.set("")  # Resetting the Combobox
    
    summary_text.configure(state='normal')
    summary_text.delete(1.0, tk.END)
    summary_text.configure(state='disabled')

def lookup_airline(airline_code, worksheet):
    for row in worksheet.iter_rows(min_col=1, max_col=2, values_only=True):
        if row[0] and row[0].lower() == airline_code:
            return row[1]
    return None

def lookup_aircraft(aircraft, worksheet):
    for row in worksheet.iter_rows(min_col=1, max_col=3, values_only=True):
        if row[0] and aircraft.lower() in row[0].lower():
            return row[1], row[2]  # tail, plane
    return None, None

def update_summary():
    airline_code = get_airline_code(entries['Ref Flight Number:'].get())
    airline = lookup_airline(airline_code, ws_airlines)
    
    tail, plane = lookup_aircraft(aircraft_dropdown.get(), ws_aircrafts)
    
    # Determine if the flight number should have 3 or 4 digits
    num_digits = random.choice([3, 4])
    if num_digits == 3:
        flight_number = random.randint(100, 999)
    else:
        flight_number = random.randint(1000, 9999)

    summary_var.set(
        f"Flight Number: {airline_code.upper()}{flight_number}\n"
        f"Airline: {airline}\n"
        f"Aircraft: {plane}\n"
        f"Tail Number: {tail}\n"
        f"Departure: {entries['Departure:'].get().upper()}\n"
        f"Destination: {entries['Destination:'].get().upper()}"
    )

    # Clear the Text widget
    summary_text.configure(state='normal')
    summary_text.delete(1.0, tk.END)
    
    # Inserting the labeled text
    summary_text.insert(tk.END, "Flight Number: ", "label")
    summary_text.insert(tk.END, f"{airline_code.upper()}{flight_number}\n", "info")

    summary_text.insert(tk.END, "Airline: ", "label")
    summary_text.insert(tk.END, f"{airline}\n", "info")

    summary_text.insert(tk.END, "Aircraft: ", "label")
    summary_text.insert(tk.END, f"{plane}\n", "info")

    summary_text.insert(tk.END, "Tail Number: ", "label")
    summary_text.insert(tk.END, f"{tail}\n", "info")

    summary_text.insert(tk.END, "Departure: ", "label")
    summary_text.insert(tk.END, f"{entries['Departure:'].get().upper()}\n", "info")

    summary_text.insert(tk.END, "Destination: ", "label")
    summary_text.insert(tk.END, f"{entries['Destination:'].get().upper()}\n", "info")

    summary_text.configure(state='disabled')

def search_airline(event=None):
    airline_code = get_airline_code(entries["Ref Flight Number:"].get())
    if not airline_code:
        airline_var.set("Invalid Flight Number")
        return

    result = lookup_airline(airline_code, ws_airlines)

    if result:
        airline_var.set(result)
    else:
        airline_var.set("Airline Not Found")

def discover_location(location_field, event=None):
    location = entries[location_field].get()
    webbrowser.open(f"https://flightsim.to/discover/{location}")

def discover_aircraft():
    airline_code = get_airline_code(entries['Ref Flight Number:'].get())
    airline = lookup_airline(airline_code, ws_airlines)
    if airline:
        airline_modified = airline.replace(" ", "%20")
        webbrowser.open(f"https://flightsim.to/discover/{airline_modified}")

def get_all_aircrafts(worksheet):
    aircrafts = []
    for row in worksheet.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):  # Starting from row 3 and getting aircraft names from column A
        aircrafts.append(row[0])
    return aircrafts

wb = openpyxl.load_workbook(get_path_from_config("Database"))
ws_airlines = wb["Airlines"]
ws_aircrafts = wb["Aircraft"]

all_aircrafts = get_all_aircrafts(ws_aircrafts)

icon_path = os.path.join(current_directory, 'MSFS.ico')

app = tk.Tk()
app.resizable(False, False)
app.title('Flight Setup')
app.iconbitmap(default=icon_path)
app.configure(bg='#1c1a27')
style = ttk.Style()
style.theme_use('alt')  
style.configure('TCombobox', background='#1c1a27', foreground='#1c1a27', fieldbackground='#7f62ca', selectbackground='#7f62ca', selectforeground='#1c1a27')
style.configure('TButton', font=default_font)

summary_var = tk.StringVar(app)

btn_pick_flight = tk.Button(app, text="Pick a flight", command=on_button_click, bg='#7f62ca', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27' , font=default_font)
btn_pick_flight.grid(row=0, column=0, columnspan=2, padx=5, pady=10, sticky="ew")

labels_text = ["Ref Flight Number:", "Departure:", "Destination:"]
entries = {}

for idx, label_text in enumerate(labels_text):
    btn = tk.Button(app,font=default_font , text=label_text, width=15, bg='#7f62ca', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27')
    
    if label_text == "Ref Flight Number:":
        btn.config(command=search_airline)
        entries[label_text] = entry = tk.Entry(app, font=default_font , width=20, bg='#7f62ca', fg='#1c1a27', insertbackground='#1c1a27')
        entry.bind('<Return>', search_airline)
    elif label_text in ["Departure:", "Destination:"]:
        btn.config(command=lambda field=label_text: discover_location(field))
        entries[label_text] = entry = tk.Entry(app, font=default_font, width=20, bg='#7f62ca', fg='#1c1a27', insertbackground='#1c1a27')
        entry.bind('<Return>', lambda event, field=label_text: discover_location(field, event))
        
    btn.grid(row=idx + 1, column=0, padx=5, pady=10, sticky="e")
    entry.grid(row=idx + 1, column=1, padx=5, pady=10, sticky="w")

aircraft_label = ttk.Label(app, font=default_font , text="Aircraft:")
aircraft_label.grid(row=idx + 2, column=0, padx=5, pady=10, sticky="e")

aircraft_dropdown = ttk.Combobox(app, font=default_font, values=all_aircrafts, width=18)
aircraft_dropdown.grid(row=idx + 2, column=1, padx=5, pady=10, sticky="w")

aircraft_button = tk.Button(app, font=default_font, text="Aircraft", command=discover_aircraft, bg='#7f62ca', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27', width=15)
aircraft_button.grid(row=idx + 2, column=0, padx=5, pady=10, sticky="e")

btn_reset = tk.Button(app, font=default_font, text="Reset", command=reset_fields, bg='#b3261e', fg='#e0e0e2', activebackground='#737373', activeforeground='#1c1a27', width=15)
btn_reset.grid(row=idx + 3, column=0, padx=5, pady=5, sticky="e")

btn_generate = tk.Button(app, font=default_font, text="Generate", command=update_summary, bg='#e0e0e2', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27', width=15)
btn_generate.grid(row=idx + 3, column=1, padx=20, pady=5, sticky="w")

summary_font = ("Space Grotesk", 14)

summary_text = tk.Text(app, font=summary_font, height=8, width=30, bg='#1c1a27', fg='#a07bff', insertbackground='#1c1a27')
summary_text.grid(row=idx + 4, column=0, columnspan=2, padx=5, pady=10)
summary_text.tag_configure("label", foreground="#a07bff")
summary_text.tag_configure("info", foreground="#e64849")
summary_text.insert(tk.INSERT, summary_var.get())
summary_text.configure(state='disabled')

buttons_info = [
    {"text": "SimBrief", "path": "https://dispatch.simbrief.com/options/new"},
    {"text": "Charts", "path": get_path_from_config("Charts").strip('"')},
    {"text": "Addons", "path": get_path_from_config("Addons").strip('"')},
    {"text": "Checklists", "path": get_path_from_config("Checklists").strip('"')}
]

for idx, info in enumerate(buttons_info, start=idx+5):
    if "http" in info["path"]:
        btn = tk.Button(app, font=default_font , text=info["text"], command=lambda path=info["path"]: webbrowser.open(path), bg='#7f62ca', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27')
    else:
        btn = tk.Button(app, font=default_font , text=info["text"], command=lambda path=info["path"]: os.startfile(path), bg='#7f62ca', fg='#1c1a27', activebackground='#737373', activeforeground='#1c1a27')
    btn.grid(row=idx, column=0, columnspan=2, padx=5, pady=10, sticky="ew")

app.mainloop()
